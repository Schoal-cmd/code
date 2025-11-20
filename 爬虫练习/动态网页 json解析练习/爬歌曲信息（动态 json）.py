import requests as re
import openpyxl

#工作簿
wb=openpyxl.Workbook()
#工作表
ws=wb.active
ws.title='lyric'
#表头
ws['A1']='歌名'
ws['B1']='所属专辑'
ws['C1']='播放时长'
ws['D1']='播放链接'

for i in range(5):
    #获取
    res_songs=re.get('https://c.y.qq.com/soso/fcgi-bin/client_search_cp?ct=24&qqmusic_ver=1298&new_json=1&remoteplace=txt.yqq.song&searchid=67820157881444874&t=0&aggr=1&cr=1&catZhida=1&lossless=0&flag_qc=0&p=str(i)&n=10&w=%E5%91%A8%E6%9D%B0%E4%BC%A6&g_tk_new_20200303=5381&g_tk=5381&loginUin=0&hostUin=0&format=json&inCharset=utf8&outCharset=utf-8&notice=0&platform=yqq.json&needNewCode=0')
    #解析
    json_songs=res_songs.json()
    #提取
    list_songs=json_songs['data']['song']['list']
    for song in list_songs:
        name=song['name']
        album=song['album']['name']
        time=song['interval']
        href='https://y.qq.com/n/yqq/song/'+song['mid']+'.html\n\n'
        ws.append([name,album,time,href])

wb.save('song.xlsx')